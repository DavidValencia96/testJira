import requests
import json
from flask import jsonify, current_app
from base64 import b64encode
import base64
import os
import time
from datetime import datetime
import csv
from requests.auth import HTTPBasicAuth
import pandas as pd
import openpyxl
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Border, Side, PatternFill

user_email = "pablo.munoz@bebolder.co"
token_user = "ATATT3xFfGF0UrsshZ9JZEMG-0eZQbBJ_GgT5-mSghYU8ustURw07LprkdngoC1iO8Y198B4b8nIePIekKQNhL4uQAQsVeIXhpWRY3GjtN_O-j9zOS_ZixxwZPdzcqdVsw_SKfox8okJwcj_57XIu3ZM0C7iwDFD3E-vnkLo6TQpfL-i_3mV6jM=07B92C26"
domain = "https://bebolder.atlassian.net"

def obtener_sprints_jira(proyecto_id):
    url = f"{domain}/rest/greenhopper/1.0/sprintquery/{proyecto_id}?includeFutureSprints=false&_=1743513788133"
    auth = HTTPBasicAuth(user_email, token_user)
    headers = {"Accept": "application/json"}

    response = requests.get(url, headers=headers, auth=auth)

    data_folder = os.path.join(current_app.root_path, 'data')
    if not os.path.exists(data_folder):
        os.makedirs(data_folder)

    json_data_sprint = os.path.join(data_folder, 'data_sprint.json')

    if os.path.exists(json_data_sprint):
        os.remove(json_data_sprint)

    if response.status_code == 200:
        json_response = response.json()
        if "sprints" in json_response:
            sprints = json_response["sprints"]
            sprint_data = []
            for sprint in sprints:
                sprint_data.append({
                    "id": sprint["id"],
                    "name": sprint["name"],
                    "state": sprint["state"],
                    "goal": sprint.get("goal", ""),
                    "sprintVersion": sprint.get("sprintVersion", "")
                })
            with open(json_data_sprint, 'w', encoding='utf-8') as json_file:
                json.dump(sprint_data, json_file, ensure_ascii=False, indent=4)
            return sprint_data 
        else:
            raise Exception("No se encontraron sprints.")
    else:
        raise Exception(f"Error al obtener datos de Jira: {response.status_code}")

def obtener_hus_de_sprint(proyecto_id, sprint_id):
    start_time = time.time()

    print(f"Iniciando la generación del reporte para el sprint {sprint_id} del proyecto {proyecto_id}")
    
    try:
        # Cargar el archivo JSON con los datos de issues
        with open("data/issues_all_sprint.json", "r", encoding="utf-8") as file:
            issues_json = json.load(file)
        print("Archivo JSON cargado exitosamente")
    except FileNotFoundError as e:
        print(f"Error al cargar el archivo: {e}")
        return {"error": "Archivo no encontrado."}, 500
    except json.JSONDecodeError as e:
        print(f"Error en el formato del JSON: {e}")
        return {"error": "Error al decodificar el JSON."}, 500

    issues_dict = {issue["key"]: issue for issue in issues_json}    
    print(f"Diccionario de issues creado con {len(issues_dict)} elementos.")

    url = f"{domain}/rest/greenhopper/1.0/rapid/charts/sprintreport?rapidViewId={proyecto_id}&sprintId={sprint_id}&_=1743432986996"
    auth_value = b64encode(f"{user_email}:{token_user}".encode()).decode()
    headers = {
        "Authorization": f"Basic {auth_value}",
        "Accept": "application/json"
    }

    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        print("Respuesta recibida de Jira.")
    except requests.exceptions.RequestException as e:
        print(f"Error en la solicitud a Jira: {e}")
        return {"error": f"Error al hacer la solicitud a Jira: {e}"}, 500

    if response.status_code == 403:
        print("Error 403: No tienes acceso a la API de Jira.")
        return {"error": "Error 403: No tienes acceso a la API de Jira. Revisa tu correo, token o permisos."}, 403
    elif response.status_code != 200:
        print(f"Error en la petición: Código {response.status_code}")
        return {"error": f"Error en la petición: Código {response.status_code}", "details": response.text}, response.status_code

    data = response.json()
    excel_file_path = f'data/issues_sprint_{proyecto_id}_{sprint_id}.xlsx'
    rows = []
    added_during_sprint = data["contents"].get("issueKeysAddedDuringSprint", {})
    total_planeado = 0
    total_ejecutado = 0

    for issue in data["contents"].get("completedIssues", []):
        if issue.get("typeName") in ["Historia", "Error"]:
            added_during_sprint_value = added_during_sprint.get(issue.get("key"), False)
            updated_at = datetime.utcfromtimestamp(issue.get("updatedAt", 0) / 1000).strftime('%Y-%m-%d %H:%M:%S')
            categoria = "Issue completado en el sprint actual"
            total_ejecutado += 1

            json_issue = issues_dict.get(issue.get("key"))
            
            if json_issue:
                customfield = json_issue["fields"].get("customfield_10048")
                definitionOfFact = customfield.get("value") if isinstance(customfield, dict) else ""
                storyPoint = json_issue["fields"].get("customfield_10033", 0.0)
                storyPointEstimated = json_issue["fields"].get("customfield_10016", 0.0)
                storyPointExecuted = json_issue["fields"].get("customfield_10046", 0.0)

                rows.append([
                    issue.get("key"), 
                    issue.get("typeName"), 
                    issue.get("summary"),
                    issue.get("priorityName"),
                    issue.get("assigneeName"),
                    issue.get("statusName"), 
                    definitionOfFact,
                    issue.get("projectId"), 
                    updated_at,
                    categoria, 
                    added_during_sprint_value,
                    storyPoint,
                    storyPointEstimated,
                    storyPointExecuted
                ])
                total_planeado += 1
            else:
                print(f"Warning: No se encontró el issue con key {issue.get('key')} en el archivo JSON.")

    for issue in data["contents"].get("issuesNotCompletedInCurrentSprint", []):
        if issue.get("typeName") in ["Historia", "Error"]:
            added_during_sprint_value = added_during_sprint.get(issue.get("key"), False)
            updated_at = datetime.utcfromtimestamp(issue.get("updatedAt", 0) / 1000).strftime('%Y-%m-%d %H:%M:%S')
            categoria = "Issue NO completado en el sprint actual"

            json_issue = issues_dict.get(issue.get("key"))

            if json_issue:
                customfield = json_issue["fields"].get("customfield_10048")
                definitionOfFact = customfield.get("value") if isinstance(customfield, dict) else ""
                storyPoint = json_issue["fields"].get("customfield_10033", 0.0)
                storyPointEstimated = json_issue["fields"].get("customfield_10016", 0.0)
                storyPointExecuted = json_issue["fields"].get("customfield_10046", 0.0)

            rows.append([
                issue.get("key"), 
                issue.get("typeName"), 
                issue.get("summary"),
                issue.get("priorityName"), 
                issue.get("assigneeName"),
                issue.get("statusName"), 
                definitionOfFact,
                issue.get("projectId"), 
                updated_at,
                categoria, 
                added_during_sprint_value,
                storyPoint,
                storyPointEstimated,
                storyPointExecuted
            ])
            total_planeado += 1

    for issue in data["contents"].get("puntedIssues", []):
        if issue.get("typeName") in ["Historia", "Error"]:
            added_during_sprint_value = added_during_sprint.get(issue.get("key"), False)
            updated_at = datetime.utcfromtimestamp(issue.get("updatedAt", 0) / 1000).strftime('%Y-%m-%d %H:%M:%S')
            categoria = "Issue postergado del sprint actual"

            json_issue = issues_dict.get(issue.get("key"))

            if json_issue:
                customfield = json_issue["fields"].get("customfield_10048")
                definitionOfFact = customfield.get("value") if isinstance(customfield, dict) else ""
                storyPoint = json_issue["fields"].get("customfield_10033", 0.0)
                storyPointEstimated = json_issue["fields"].get("customfield_10016", 0.0)
                storyPointExecuted = json_issue["fields"].get("customfield_10046", 0.0)

            rows.append([
                issue.get("key"), 
                issue.get("typeName"), 
                issue.get("summary"),
                issue.get("priorityName"), 
                issue.get("assigneeName"),
                issue.get("statusName"),
                definitionOfFact,
                issue.get("projectId"), 
                updated_at,
                categoria,
                added_during_sprint_value,
                storyPoint,
                storyPointEstimated,
                storyPointExecuted
            ])
            total_planeado += 1

    porcentaje_cumplimiento = (total_ejecutado / total_planeado) * 100 if total_planeado > 0 else 0

    if rows:
        df = pd.DataFrame(rows, columns=[
            "key", "Tipo issue", "Titulo", "Prioridad", "Assignado", 
            "Estado", "Definicion de hecho", "ID Proyecto", "Actualizada", "Categoria", "Añadido durante el sprint",
            "Puntos de historia", "Puntos estimados", "Puntos ejecutados"
        ])
        df.to_excel(excel_file_path, index=False, sheet_name="Reporte Sprint")
        print(f"Archivo Excel generado en {excel_file_path}")
        wb = openpyxl.load_workbook(excel_file_path)
        summary_sheet = wb.create_sheet("Resumen Sprint")

        summary_sheet["A1"] = "Velocidad"
        summary_sheet["A2"] = "Planeado"
        summary_sheet["B2"] = total_planeado
        summary_sheet["A3"] = "Ejecutado"
        summary_sheet["B3"] = total_ejecutado
        summary_sheet["A4"] = "% Cumplimiento"
        summary_sheet["B4"] = f"{porcentaje_cumplimiento:.2f}%"

        # ----- Resumen exclusivo de historias -----
        historias_df = df[df["Tipo issue"] == "Historia"]
        total_historias = len(historias_df)
        historias_adicionadas = historias_df["Añadido durante el sprint"].sum()
        historias_ejecutadas = len(historias_df[historias_df["Categoria"] == "Issue completado en el sprint actual"])
        porcentaje_historias = (historias_ejecutadas / total_historias) * 100 if total_historias > 0 else 0

        summary_sheet["A7"] = "Historias"
        summary_sheet["A8"] = "Planeado"
        summary_sheet["B8"] = total_historias
        summary_sheet["A9"] = "H. Adicionadas"
        summary_sheet["B9"] = historias_adicionadas
        summary_sheet["A10"] = "Ejecutado"
        summary_sheet["B10"] = historias_ejecutadas
        summary_sheet["A11"] = "% Cumplimiento"
        summary_sheet["B11"] = f"{porcentaje_historias:.2f}%"

        # ----- Resumen exclusivo de errores -----
        errores_df = df[df["Tipo issue"] == "Error"]
        total_errores = len(errores_df)
        errores_cerrados = len(errores_df[errores_df["Estado"] == "HU en producción"])
        errores_activos = total_errores - errores_cerrados
        porcentaje_errores = (errores_cerrados / total_errores) * 100 if total_errores > 0 else 0

        summary_sheet["A13"] = "Errores (BUGS)"
        summary_sheet["A14"] = "Reportados"
        summary_sheet["B14"] = total_errores
        summary_sheet["A15"] = "Cerrados"
        summary_sheet["B15"] = errores_cerrados
        summary_sheet["A16"] = "Activos"
        summary_sheet["B16"] = errores_activos
        summary_sheet["A17"] = "% Cumplimiento"
        summary_sheet["B17"] = f"{porcentaje_errores:.2f}%"

        # ----- Resumen de Añadidos durante el sprint -----
        añadidos_df = df[df["Añadido durante el sprint"] == True]
        historias_añadidas = len(añadidos_df[añadidos_df["Tipo issue"] == "Historia"])
        errores_añadidos = len(añadidos_df[añadidos_df["Tipo issue"] == "Error"])

        summary_sheet["A19"] = "Añadidos"
        summary_sheet["A20"] = "Historias"
        summary_sheet["B20"] = historias_añadidas
        summary_sheet["A21"] = "Errores"
        summary_sheet["B21"] = errores_añadidos

        # ----- Estilos-----
        borde_negro = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        relleno_cabecera = PatternFill(start_color="E36205", end_color="E36205", fill_type="solid")

        for fila in summary_sheet.iter_rows(min_row=1, max_row=21, min_col=1, max_col=2):
            for celda in fila:
                if celda.value is not None:
                    celda.border = borde_negro

        summary_sheet["A1"].fill = relleno_cabecera
        summary_sheet["A7"].fill = relleno_cabecera
        summary_sheet["A13"].fill = relleno_cabecera
        summary_sheet["A19"].fill = relleno_cabecera

        wb.save(excel_file_path)
        print("Archivo Excel guardado exitosamente.")
    else:
        print("No se encontraron issues para incluir en el reporte.")

    if not os.path.exists('data'):
        os.makedirs('data')

    end_time = time.time()
    execution_time = end_time - start_time
    print(f"Tiempo de ejecución: {execution_time} segundos")

    return {"archivo": f"https://testjira.onrender.com/{excel_file_path}", "tiempo": execution_time}

def generar_json_issues(proyecto_key):
    issue_types = ["Bug", "Story"]

    auth_string = f"{user_email}:{token_user}"
    encoded_auth = base64.b64encode(auth_string.encode()).decode()

    jql_query = f"project={proyecto_key}"
    if issue_types:
        tipos = ', '.join([f'"{tipo}"' for tipo in issue_types])
        jql_query += f" AND type IN ({tipos})"

    archivo_json = "data/issues_all_sprint.json"

    if os.path.exists(archivo_json):
        os.remove(archivo_json)
        print(f"Archivo {archivo_json} eliminado.")

    with open(archivo_json, "w", encoding="utf-8") as json_file:
        json.dump([], json_file, indent=4, ensure_ascii=False)

    start_at = 0
    max_results = 100
    total = 0

    start_time = time.time()
    success = False

    try:
        with open(archivo_json, "r+", encoding="utf-8") as json_file:
            while True:
                url = f"{domain}/rest/api/3/search?jql={jql_query}&startAt={start_at}&maxResults={max_results}"
                headers = {
                    "Authorization": f"Basic {encoded_auth}",
                    "Accept": "application/json"
                }

                try:
                    response = requests.get(url, headers=headers, timeout=130)
                    response.raise_for_status()
                    json_response = response.json()

                    if total == 0:
                        total = json_response['total']

                    json_file.seek(0)
                    try:
                        current_data = json.load(json_file)
                    except json.JSONDecodeError:
                        current_data = []

                    if isinstance(json_response.get('issues', []), list):
                        current_data.extend(json_response['issues'])
                    else:
                        print(f"Error: Los issues no están en el formato esperado. Tipo de datos: {type(json_response.get('issues'))}")

                    json_file.seek(0)
                    json.dump(current_data, json_file, indent=4, ensure_ascii=False)
                    print(f"Datos guardados exitosamente para el rango de startAt {start_at}")

                    start_at += max_results

                    if start_at >= total:
                        success = True
                        break

                except requests.exceptions.Timeout:
                    print("La solicitud ha excedido el tiempo de espera.")
                    return jsonify({"error": "La solicitud a Jira ha excedido el tiempo de espera."}), 408

                except requests.exceptions.RequestException as e:
                    print(f"Ocurrió un error en la solicitud: {e}")
                    return jsonify({"error": f"Error en la extracción de datos: {e}"}), 500

    except Exception as e:
        print(f"Ocurrió un error al intentar abrir o escribir en el archivo: {e}")
        return jsonify({"error": "Error al manejar el archivo JSON."}), 500

    end_time = time.time()
    execution_time = round(end_time - start_time, 2)

    if success:
        return {
            "archivo": f"https://testjira.onrender.com/{archivo_json}", "tiempo": execution_time
        }
    else:
        return jsonify({"error": "Hubo un problema al procesar los issues."}), 500
